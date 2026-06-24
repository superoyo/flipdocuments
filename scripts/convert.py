#!/usr/bin/env python3
"""
Convert a weekly TV media-schedule grid (a "Pulzar-style" plan) into a monthly
channel grid (a "Kirei-style" plan).

INPUT  (Pulzar-style): a weekly calendar where each weekday is a group of 6
  columns [CH, <WEEKDAY>, TVR, Mat., Cost/Spot, Time]. The 7 weekday groups sit
  side by side. Spots are listed vertically under each day; the date number for
  each day sits in that group's Time column on a "date row". Each spot cell reads
  like "เรื่องเล่าเช้านี้: 30x1" (program name : <seconds>x<spots>). The material
  version (A / B / C ...) is in the Mat. column.

OUTPUT (Kirei-style): one row per (channel, program, duration); dates 1..N of the
  month as columns; each day cell holds the version code(s) that ran that day
  (e.g. "A", "A2", "BC"); right-hand columns show สปอต / ราคาต่อหน่วย / % ลด /
  รวมเงินสุทธิ. Channels are grouped with a TOTAL CHANNEL row, then a grand
  สรุปรวมทั้งสิ้น row, a 10% service-fee footer, and an auto-detected A/B/C legend.

Usage:
    python convert.py INPUT.xlsx [OUTPUT.xlsx] [--month "พ.ค. 2569"]
        [--client "..."] [--product "..."] [--campaign "..."] [--docno "..."]

The parser AUTO-DETECTS the header row, the weekday column groups, the date rows
and the legend, so it works for any month / any layout that follows the same
column convention — nothing about rows or dates is hard-coded.
"""
import sys, re, argparse, datetime
import openpyxl
from openpyxl.utils import get_column_letter as gl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.properties import PageSetupProperties
from collections import defaultdict

WEEKDAYS = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']
WD_TH = ['จ', 'อ', 'พ', 'พฤ', 'ศ', 'ส', 'อา']
MONTHS = {  # english/abbrev -> (thai abbrev, month number)
    'jan': ('ม.ค.', 1), 'feb': ('ก.พ.', 2), 'mar': ('มี.ค.', 3), 'apr': ('เม.ย.', 4),
    'may': ('พ.ค.', 5), 'jun': ('มิ.ย.', 6), 'jul': ('ก.ค.', 7), 'aug': ('ส.ค.', 8),
    'sep': ('ก.ย.', 9), 'oct': ('ต.ค.', 10), 'nov': ('พ.ย.', 11), 'dec': ('ธ.ค.', 12),
}
DAYS_IN_MONTH = {1: 31, 2: 29, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31}
PROG_RE = re.compile(r'^(.*?):\s*(\d+(?:\.\d+)?)\s*x\s*(\d+)\s*$')
LEGEND_TAG_RE = re.compile(r'^([A-E])\s*:?\s*$')        # the version tag in the CH column
SEC_RE = re.compile(r'(\d+(?:\.\d+)?)\s*sec\b', re.IGNORECASE)  # find duration inside a creative name


def norm(v):
    return str(v).strip() if v is not None else ''


def tnum(t):
    s = norm(t)
    return int(float(s)) if s.replace('.', '').replace('-', '').isdigit() and '-' not in s else None


def fmt_time(t):
    n = tnum(t)
    return f"{n // 100:02d}:{n % 100:02d}" if n is not None and n >= 100 else norm(t)


# --------------------------------------------------------------------------- #
# 1. PARSE the source workbook
# --------------------------------------------------------------------------- #
def find_sheet_and_header(wb):
    """Locate the sheet + header row containing the CH / Mat. / Cost/Spot / Time
    columns, and return the weekday-group column map."""
    for ws in wb.worksheets:
        maxr, maxc = min(ws.max_row, 80), min(ws.max_column, 200)
        for r in range(1, maxr + 1):
            vals = {norm(ws.cell(r, c).value).upper(): c for c in range(1, maxc + 1)
                    if ws.cell(r, c).value is not None}
            wk_cols = {v: c for v, c in vals.items() if v in WEEKDAYS}
            if len(wk_cols) >= 3 and any(k.startswith('MAT') for k in vals) and \
               any('COST' in k for k in vals):
                groups = {}
                for wd, wcol in wk_cols.items():
                    # group layout: CH(W-1) WD(W) TVR(W+1) Mat(W+2) Cost(W+3) Time(W+4)
                    groups[wd] = dict(ch=wcol - 1, prog=wcol, tvr=wcol + 1,
                                      mat=wcol + 2, cost=wcol + 3, time=wcol + 4)
                return ws, r, groups
    raise SystemExit("Could not find a Pulzar-style header (need CH / weekday / Mat. / Cost/Spot / Time).")


def parse(ws, header_row, groups):
    """Walk the rows, tracking the current date per weekday column, and collect
    one record per spot."""
    rows = defaultdict(lambda: dict(days=defaultdict(lambda: defaultdict(int)),
                                    cost=set(), times=defaultdict(int), wds=set()))
    cur_date = {wd: None for wd in groups}
    date_wd = {}            # date number -> weekday index (for weekend shading)
    legend = {}             # version letter -> creative description
    month_txt = None
    dur_of_letter = {}

    for r in range(header_row + 1, ws.max_row + 1):
        for wd, g in groups.items():
            # (a) date marker?
            tv = ws.cell(r, g['time']).value
            n = tnum(tv)
            prog = norm(ws.cell(r, g['prog']).value)
            # (b) month name often sits in a program cell on a date row
            if month_txt is None and prog[:3].lower() in MONTHS:
                month_txt = prog[:3].lower()
            if n is not None and 1 <= n <= 31 and not PROG_RE.match(prog):
                cur_date[wd] = n
                date_wd[n] = WEEKDAYS.index(wd)
                continue
            if not prog:
                continue
            # (c) legend row?  CH column holds a version tag (A / B: / C:) and the
            #     program column holds the creative name, e.g. "Truck 15 sec พี่หนุ่ม"
            tag = LEGEND_TAG_RE.match(norm(ws.cell(r, g['ch']).value))
            if tag and not PROG_RE.match(prog):
                letter = tag.group(1)
                if letter not in legend:
                    legend[letter] = prog
                    sm = SEC_RE.search(prog)
                    if sm:
                        dur_of_letter[letter] = int(float(sm.group(1)))
                continue
            # (d) a real spot
            m = PROG_RE.match(prog)
            if not m or cur_date[wd] is None:
                continue
            name = m.group(1).strip()
            dur = int(float(m.group(2)))
            spots = int(m.group(3))
            mat = norm(ws.cell(r, g['mat']).value) or '?'
            cost = ws.cell(r, g['cost']).value
            ch = norm(ws.cell(r, g['ch']).value)
            o = rows[(ch, name, dur)]
            o['days'][cur_date[wd]][mat] += spots
            if cost not in (None, ''):
                try:
                    o['cost'].add(round(float(cost)))
                except (TypeError, ValueError):
                    pass
            o['times'][norm(ws.cell(r, g['time']).value)] += 1
            o['wds'].add(WEEKDAYS.index(wd))
            dur_of_letter.setdefault(mat, dur)

    return rows, date_wd, legend, dur_of_letter, month_txt


# --------------------------------------------------------------------------- #
# 2. helpers for output formatting
# --------------------------------------------------------------------------- #
def day_pattern(wds):
    s = sorted(wds)
    if not s:
        return ''
    if s == list(range(s[0], s[-1] + 1)):
        return WD_TH[s[0]] if len(s) == 1 else f"{WD_TH[s[0]]}-{WD_TH[s[-1]]}"
    return '/'.join(WD_TH[i] for i in s)


def rep_time(times):
    t = max(times, key=lambda k: times[k]) if times else ''
    return fmt_time(t)


def cell_code(mats):
    """Render a day's versions: 'A', 'A2', 'BC', 'B2C' ..."""
    out = ''
    for L in sorted(mats):
        c = mats[L]
        out += L + (str(c) if c > 1 else '')
    return out


def weekday_filler(date_wd, ndays):
    """Given some known date->weekday pairs, infer weekday for every 1..ndays."""
    if date_wd:
        ref_d, ref_wd = next(iter(date_wd.items()))
        return [((ref_wd + (d - ref_d)) % 7) for d in range(1, ndays + 1)]
    return [((4 + (d - 1)) % 7) for d in range(1, ndays + 1)]  # fallback: 1st = Fri


# --------------------------------------------------------------------------- #
# 3. BUILD the output workbook
# --------------------------------------------------------------------------- #
def build(rows, date_wd, legend, dur_of_letter, month_txt, meta, out_path):
    F = 'Tahoma'
    NUM = '#,##0;(#,##0);"-"'
    thin = Side(style='thin', color='B0B0B0')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    HEAD = PatternFill('solid', fgColor='1F4E78')
    CHFILL = PatternFill('solid', fgColor='2E75B6')
    TOT = PatternFill('solid', fgColor='DDEBF7')
    WE = PatternFill('solid', fgColor='FCE4D6')
    WEHEAD = PatternFill('solid', fgColor='C55A11')

    th, mn = (MONTHS[month_txt][0], MONTHS[month_txt][1]) if month_txt in MONTHS else ('', 0)
    ndays = DAYS_IN_MONTH.get(mn, 31)
    month_label = meta.get('month') or (f"{th} {meta.get('year', '')}".strip() if th else '')
    wd_by_day = weekday_filler(date_wd, ndays)

    DAY0 = 5
    LAST = DAY0 + ndays - 1
    C_SPOT, C_RATE, C_DISC, C_NET = LAST + 1, LAST + 2, LAST + 3, LAST + 4
    NC = C_NET

    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = 'Media Grid'

    def put(r, c, v, sz=11, b=False, i=False, fc='000000', h='left', va='center',
            fill=None, nf=None, border=True, wrap=False):
        cell = sh.cell(r, c, v)
        cell.font = Font(F, size=sz, bold=b, italic=i, color=fc)
        cell.alignment = Alignment(h, va, wrap_text=wrap)
        if fill:
            cell.fill = fill
        if nf:
            cell.number_format = nf
        if border:
            cell.border = bd
        return cell

    # column widths
    for col, w in [('A', 30), ('B', 7), ('C', 6.5), ('D', 7)]:
        sh.column_dimensions[col].width = w
    for d in range(ndays):
        sh.column_dimensions[gl(DAY0 + d)].width = 3.6
    for c, w in [(C_SPOT, 6), (C_RATE, 11), (C_DISC, 6), (C_NET, 14)]:
        sh.column_dimensions[gl(c)].width = w

    # ---- header block ----
    sh.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NC)
    put(1, 1, 'TELEVISION  ตารางการใช้สื่อ', sz=16, b=True, fc='FFFFFF', h='center', fill=HEAD, border=False)
    sh.row_dimensions[1].height = 26
    put(2, 1, 'ลูกค้า', b=True, border=False);    put(2, 2, ':  ' + meta.get('client', ''), border=False)
    put(3, 1, 'ผลิตภัณฑ์', b=True, border=False); put(3, 2, ':  ' + meta.get('product', ''), border=False)
    put(4, 1, 'แคมเปญ', b=True, border=False);    put(4, 2, ':  ' + meta.get('campaign', ''), border=False)
    put(2, C_DISC, 'ระยะเวลาโฆษณา', b=True, h='right', border=False); put(2, C_NET, ':  ' + month_label, border=False)
    put(3, C_DISC, 'เอกสารเลขที่', b=True, h='right', border=False);  put(3, C_NET, ':  ' + meta.get('docno', ''), border=False)
    put(4, C_DISC, 'วันที่', b=True, h='right', border=False);        put(4, C_NET, ':  ' + datetime.date.today().strftime('%d/%m/%Y'), border=False)

    HR1, HR2, DATA0 = 6, 7, 8
    for col, lab in [(1, 'รายการ'), (2, 'วัน'), (3, 'วินาที'), (4, 'เวลา'),
                     (C_SPOT, 'สปอต'), (C_RATE, 'ราคาต่อหน่วย'), (C_DISC, '% ลด'), (C_NET, 'รวมเงินสุทธิ\nบาท')]:
        sh.merge_cells(start_row=HR1, start_column=col, end_row=HR2, end_column=col)
        put(HR1, col, lab, sz=10, b=True, fc='FFFFFF', h='center', fill=HEAD, wrap=True)
    for d in range(1, ndays + 1):
        cc = DAY0 + d - 1
        we = wd_by_day[d - 1] in (5, 6)
        put(HR1, cc, WD_TH[wd_by_day[d - 1]], sz=8, b=True, fc='FFFFFF', h='center', fill=(WEHEAD if we else HEAD))
        put(HR2, cc, d, sz=9, b=True, fc=('000000' if we else 'FFFFFF'), h='center', fill=(WE if we else CHFILL))
    sh.row_dimensions[HR1].height = 16
    sh.row_dimensions[HR2].height = 16

    # ---- channel ordering by total spend (desc) ----
    chan_money = defaultdict(int)
    for (ch, nm, dur), o in rows.items():
        cost = max(o['cost']) if o['cost'] else 0
        spots = sum(c for dd in o['days'].values() for c in dd.values())
        chan_money[ch] += cost * spots
    chan_order = sorted(chan_money, key=lambda c: -chan_money[c])

    row = DATA0
    chan_total_rows = []
    day_grand = defaultdict(int)

    for ch in chan_order:
        sh.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        put(row, 1, ch, b=True, fc='FFFFFF', fill=CHFILL)
        for c in range(5, NC + 1):
            put(row, c, None, fill=CHFILL)
        sh.row_dimensions[row].height = 18
        row += 1

        items = [(k, o) for k, o in rows.items() if k[0] == ch]
        items.sort(key=lambda x: (min((tnum(t) or 999999) for t in x[1]['times']), -x[0][2], x[0][1]))
        prog_rows = []
        day_chan = defaultdict(int)
        for (c2, name, dur), o in items:
            cost = max(o['cost']) if o['cost'] else 0
            spots = sum(c for dd in o['days'].values() for c in dd.values())
            put(row, 1, ' ' + name)
            put(row, 2, day_pattern(o['wds']), sz=10, h='center')
            put(row, 3, dur, sz=10, h='center')
            put(row, 4, rep_time(o['times']), sz=10, h='center')
            for d in range(1, ndays + 1):
                cc = DAY0 + d - 1
                we = wd_by_day[d - 1] in (5, 6)
                mats = o['days'].get(d)
                put(row, cc, cell_code(mats) if mats else None, sz=9, h='center', fill=(WE if we else None))
                if mats:
                    cnt = sum(mats.values())
                    day_chan[d] += cnt
                    day_grand[d] += cnt
            put(row, C_SPOT, spots, sz=10, b=True, h='center')
            put(row, C_RATE, cost, sz=10, h='right', nf=NUM)
            put(row, C_DISC, None, h='center')
            put(row, C_NET, f'={gl(C_SPOT)}{row}*{gl(C_RATE)}{row}', sz=10, h='right', nf=NUM)
            prog_rows.append(row)
            row += 1

        put(row, 1, 'TOTAL CHANNEL', sz=10, b=True, fill=TOT)
        for c in (2, 3, 4):
            put(row, c, None, fill=TOT)
        for d in range(1, ndays + 1):
            cc = DAY0 + d - 1
            we = wd_by_day[d - 1] in (5, 6)
            v = day_chan.get(d)
            put(row, cc, v if v else None, sz=9, b=True, h='center', fill=(WE if we else TOT))
        put(row, C_SPOT, f'=SUM({gl(C_SPOT)}{prog_rows[0]}:{gl(C_SPOT)}{prog_rows[-1]})', sz=10, b=True, h='center', fill=TOT)
        put(row, C_RATE, None, fill=TOT)
        put(row, C_DISC, None, fill=TOT)
        put(row, C_NET, f'=SUM({gl(C_NET)}{prog_rows[0]}:{gl(C_NET)}{prog_rows[-1]})', sz=10, b=True, h='right', fill=TOT, nf=NUM)
        chan_total_rows.append(row)
        sh.row_dimensions[row].height = 17
        row += 2

    # ---- grand summary ----
    put(row, 1, 'สรุปรวมทั้งสิ้น', sz=12, b=True, fc='FFFFFF', fill=HEAD)
    for c in (2, 3, 4):
        put(row, c, None, fill=HEAD)
    for d in range(1, ndays + 1):
        cc = DAY0 + d - 1
        we = wd_by_day[d - 1] in (5, 6)
        v = day_grand.get(d)
        put(row, cc, v if v else None, sz=9, b=True, fc=('000000' if we else 'FFFFFF'), h='center', fill=(WE if we else HEAD))
    put(row, C_SPOT, '=' + '+'.join(f'{gl(C_SPOT)}{t}' for t in chan_total_rows), sz=11, b=True, fc='FFFFFF', h='center', fill=HEAD)
    put(row, C_RATE, None, fill=HEAD)
    put(row, C_DISC, None, fill=HEAD)
    put(row, C_NET, '=' + '+'.join(f'{gl(C_NET)}{t}' for t in chan_total_rows), sz=11, b=True, fc='FFFFFF', h='right', fill=HEAD, nf=NUM)
    grand_row = row
    sh.row_dimensions[row].height = 22
    row += 2

    # ---- money footer ----
    def foot(label, formula, bold=False):
        nonlocal row
        sh.merge_cells(start_row=row, start_column=C_SPOT, end_row=row, end_column=C_DISC)
        put(row, C_SPOT, label, h='right', b=bold, border=False)
        put(row, C_NET, formula, h='right', b=bold, nf=NUM, border=False)
        row += 1
    foot('รวมเงินสุทธิ', f'={gl(C_NET)}{grand_row}')
    foot('ค่าบริการ ร้อยละ 10', f'={gl(C_NET)}{grand_row}*0.1')
    foot('จำนวนเงินรวมทั้งสิ้น', f'={gl(C_NET)}{grand_row}*1.1', bold=True)
    sh.merge_cells(start_row=row, start_column=C_SPOT, end_row=row, end_column=C_NET)
    put(row, C_SPOT, '(ไม่รวมภาษีมูลค่าเพิ่ม)', i=True, sz=10, h='right', border=False)
    row += 2

    # ---- legend (auto-detected) ----
    put(row, 1, 'หมายเหตุ', b=True, border=False)
    row += 1
    for L in sorted(legend):
        put(row, 1, f'({L})   {legend[L]}', sz=10, border=False)
        row += 1
    put(row, 1, 'ตัวเลขหลังอักษร = จำนวนสปอต/วัน เช่น A4 = 4 สปอต, BC = B และ C อย่างละ 1 สปอต',
        i=True, sz=9, fc='595959', border=False)

    # ---- print setup ----
    sh.freeze_panes = sh.cell(DATA0, 5)
    sh.sheet_view.showGridLines = False
    sh.page_setup.orientation = 'landscape'
    sh.page_setup.paperSize = sh.PAPERSIZE_A4
    sh.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sh.page_setup.fitToWidth = 1
    sh.page_setup.fitToHeight = 0
    sh.print_title_rows = f'{HR1}:{HR2}'
    for m in ('left', 'right', 'top', 'bottom'):
        setattr(sh.page_margins, m, 0.3)

    wb.save(out_path)
    grand_spots = sum(day_grand.values())
    grand_money = sum((max(o['cost']) if o['cost'] else 0) *
                      sum(c for dd in o['days'].values() for c in dd.values())
                      for o in rows.values())
    return dict(rows=len(rows), channels=len(chan_order), spots=grand_spots,
                money=grand_money, ndays=ndays, month=month_label)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('input')
    ap.add_argument('output', nargs='?')
    ap.add_argument('--month'); ap.add_argument('--year', default='2569')
    ap.add_argument('--client', default=''); ap.add_argument('--product', default='')
    ap.add_argument('--campaign', default=''); ap.add_argument('--docno', default='')
    a = ap.parse_args()
    out = a.output or re.sub(r'\.xlsx$', '', a.input) + '_Media Grid.xlsx'

    wb = openpyxl.load_workbook(a.input, data_only=True)
    ws, hdr, groups = find_sheet_and_header(wb)
    rows, date_wd, legend, dur_letter, month_txt = parse(ws, hdr, groups)
    if not rows:
        raise SystemExit("No spots parsed — check that the file follows the Pulzar-style layout.")
    meta = dict(month=a.month, year=a.year, client=a.client,
                product=a.product, campaign=a.campaign, docno=a.docno)
    info = build(rows, date_wd, legend, dur_letter, month_txt, meta, out)
    print(f"OK  rows={info['rows']} channels={info['channels']} "
          f"spots={info['spots']} net={info['money']:,} "
          f"days={info['ndays']} month='{info['month']}'")
    print(f"Saved: {out}")


if __name__ == '__main__':
    main()
